import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import threading
import time
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font

# Variáveis globais de controle
parar_flag = threading.Event()
linhas_processadas = 0
total_duplicatas = 0
log_duplicatas = []

def filtrar_duplicatas_cpf_data(df):
    """
    Filtra duplicatas baseado em CPF, data_vencimento e numero_prestacao
    Aplica as seguintes regras em ordem de prioridade:
    
    1. Se tiver data_pagamento: salva somente a linha que NÃO tem data_pagamento
    2. Se tiver data_pagamento em ambos: traz onde cod_acordo está zerado
    3. Se tiver data_pagamento e cod_acordo em ambos: traz onde cod_prestacao é mais antigo (menor)
    4. Se data_pagamento for null: traz onde cod_acordo está zerado
    5. Se data_pagamento for null e tiver cod_acordo em ambos: traz cod_prestacao mais antiga (menor)
    """
    global total_duplicatas, log_duplicatas
    
    print(f"📊 Analisando duplicatas com nova lógica...")
    print(f"   📈 Total de registros: {len(df)}")
    
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['cpf', 'data_vencimento', 'numero_prestacao', 'cod_prestacao']
    colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]
    
    # Verificar colunas opcionais para regras de negócio
    colunas_opcionais = ['data_pagamento', 'cod_acordo']
    colunas_disponiveis = [col for col in colunas_opcionais if col in df.columns]
    
    if colunas_faltantes:
        raise ValueError(f"❌ Colunas não encontradas no arquivo: {', '.join(colunas_faltantes)}")
    
    print(f"   📋 Colunas para regras de negócio encontradas: {colunas_disponiveis}")
    
    # Identificar grupos duplicados
    grupos = df.groupby(['cpf', 'data_vencimento', 'numero_prestacao'])
    
    # Manter apenas grupos que têm mais de 1 registro (duplicados)
    grupos_duplicados = grupos.filter(lambda x: len(x) > 1)
    
    if len(grupos_duplicados) == 0:
        print(f"   ✅ Nenhum duplicado encontrado!")
        return grupos_duplicados
    
    print(f"   🔍 Duplicados encontrados: {len(grupos_duplicados)}")
    
    # Processar cada grupo de duplicatas aplicando as regras
    df_resultado = pd.DataFrame()
    grupos_processados = grupos_duplicados.groupby(['cpf', 'data_vencimento', 'numero_prestacao'])
    
    for (cpf, data_venc, num_prest), grupo in grupos_processados:
        print(f"   🎯 Processando grupo: CPF={cpf}, Data={data_venc}, Num={num_prest} ({len(grupo)} registros)")
        
        # Aplicar as regras de prioridade
        registro_escolhido = aplicar_regras_duplicatas(grupo, cpf, data_venc, num_prest)
        
        if registro_escolhido is not None:
            df_resultado = pd.concat([df_resultado, registro_escolhido], ignore_index=True)
    
    # Marcar todos os registros como "corretos" (serão salvos)
    df_resultado = df_resultado.copy()
    df_resultado['eh_menor_cod'] = True  # Todos são considerados corretos pela nova lógica
    
    total_duplicatas = len(df_resultado)
    
    print(f"   ✅ Total de registros após aplicar regras: {total_duplicatas}")
    
    # Log dos grupos processados
    grupos_unicos = df_resultado.groupby(['cpf', 'data_vencimento', 'numero_prestacao']).size()
    print(f"   📊 Grupos processados: {len(grupos_unicos)}")
    
    for (cpf, data_venc, num_prest), count in grupos_unicos.items():
        log_entry = f"Grupo resolvido: CPF={cpf}, Data Venc={data_venc}, Num Prest={num_prest} - {count} registro(s) selecionado(s)"
        log_duplicatas.append(log_entry)
        print(f"   📝 {log_entry}")
    
    return df_resultado

def aplicar_regras_duplicatas(grupo, cpf, data_venc, num_prest):
    """
    Aplica as regras específicas para escolher qual registro manter do grupo de duplicatas
    Melhorada para cobrir cenários extremos e empates
    """
    # Fazer cópia para evitar warnings
    grupo = grupo.copy()
    
    # Verificar se tem coluna data_pagamento
    tem_data_pagamento = 'data_pagamento' in grupo.columns
    tem_cod_acordo = 'cod_acordo' in grupo.columns
    
    print(f"      📋 Analisando {len(grupo)} registros do grupo")
    
    if tem_data_pagamento:
        # Verificar quais registros têm data_pagamento preenchida (incluindo valores especiais)
        tem_pagamento = (
            grupo['data_pagamento'].notna() & 
            (grupo['data_pagamento'] != '') & 
            (grupo['data_pagamento'] != 0) &
            (grupo['data_pagamento'] != '0') &
            (grupo['data_pagamento'] != 'N/A') &
            (grupo['data_pagamento'] != '0000-00-00') &
            (grupo['data_pagamento'] != '1900-01-01')
        )
        registros_com_pagamento = grupo[tem_pagamento]
        registros_sem_pagamento = grupo[~tem_pagamento]
        
        print(f"      💰 Registros COM data_pagamento: {len(registros_com_pagamento)}")
        print(f"      📅 Registros SEM data_pagamento: {len(registros_sem_pagamento)}")
        
        # REGRA 1: Se tiver data_pagamento, salva somente a linha que NÃO tem data_pagamento
        if len(registros_com_pagamento) > 0 and len(registros_sem_pagamento) > 0:
            print(f"      ✅ REGRA 1: Escolhendo registro SEM data_pagamento")
            # Se há múltiplos sem pagamento, aplicar critérios de desempate
            if len(registros_sem_pagamento) > 1:
                return aplicar_criterios_desempate(registros_sem_pagamento, "REGRA 1", tem_cod_acordo)
            return registros_sem_pagamento.iloc[[0]]
        
        # REGRA 2: Se tiver data_pagamento em ambos, traz onde cod_acordo está zerado
        elif len(registros_com_pagamento) > 0 and len(registros_sem_pagamento) == 0:
            print(f"      💰 Todos têm data_pagamento, aplicando REGRA 2")
            if tem_cod_acordo:
                # Normalizar cod_acordo para garantir comparação correta
                registros_com_pagamento = registros_com_pagamento.copy()
                registros_com_pagamento['cod_acordo_norm'] = pd.to_numeric(registros_com_pagamento['cod_acordo'], errors='coerce').fillna(-1)
                registros_cod_zero = registros_com_pagamento[registros_com_pagamento['cod_acordo_norm'] == 0]
                
                if len(registros_cod_zero) > 0:
                    print(f"      ✅ REGRA 2: Escolhendo registro com cod_acordo=0")
                    # Se há múltiplos com cod_acordo=0, aplicar critérios de desempate
                    if len(registros_cod_zero) > 1:
                        return aplicar_criterios_desempate(registros_cod_zero, "REGRA 2", tem_cod_acordo)
                    return registros_cod_zero.iloc[[0]]
                else:
                    # REGRA 3: Se todos têm data_pagamento e cod_acordo, pega menor cod_prestacao
                    print(f"      ✅ REGRA 3: Escolhendo menor cod_prestacao")
                    return aplicar_criterios_desempate(registros_com_pagamento, "REGRA 3", tem_cod_acordo)
            else:
                # Sem cod_acordo, pega menor cod_prestacao
                print(f"      ✅ Sem cod_acordo, escolhendo menor cod_prestacao")
                return aplicar_criterios_desempate(registros_com_pagamento, "SEM COD_ACORDO", tem_cod_acordo)
        
        # Se só tem registros sem data_pagamento, continua para regras 4 e 5
        else:
            grupo_sem_pagamento = registros_sem_pagamento
    else:
        # Não tem coluna data_pagamento, considera todos como sem data_pagamento
        grupo_sem_pagamento = grupo
        print(f"      📅 Coluna data_pagamento não encontrada, processando como NULL")
    
    # REGRA 4: Se data_pagamento for null, traz onde cod_acordo está zerado
    if tem_cod_acordo:
        # Normalizar cod_acordo para garantir comparação correta
        grupo_sem_pagamento = grupo_sem_pagamento.copy()
        grupo_sem_pagamento['cod_acordo_norm'] = pd.to_numeric(grupo_sem_pagamento['cod_acordo'], errors='coerce').fillna(-1)
        registros_cod_zero = grupo_sem_pagamento[grupo_sem_pagamento['cod_acordo_norm'] == 0]
        
        if len(registros_cod_zero) > 0:
            print(f"      ✅ REGRA 4: Escolhendo registro com cod_acordo=0 (data_pagamento NULL)")
            # Se há múltiplos com cod_acordo=0, aplicar critérios de desempate
            if len(registros_cod_zero) > 1:
                return aplicar_criterios_desempate(registros_cod_zero, "REGRA 4", tem_cod_acordo)
            return registros_cod_zero.iloc[[0]]
        else:
            # REGRA 5: Se data_pagamento null e tem cod_acordo em todos, pega menor cod_prestacao
            print(f"      ✅ REGRA 5: Escolhendo menor cod_prestacao (data_pagamento NULL)")
            return aplicar_criterios_desempate(grupo_sem_pagamento, "REGRA 5", tem_cod_acordo)
    else:
        # Sem cod_acordo, pega menor cod_prestacao
        print(f"      ✅ Sem cod_acordo, escolhendo menor cod_prestacao (data_pagamento NULL)")
        return aplicar_criterios_desempate(grupo_sem_pagamento, "SEM COD_ACORDO", tem_cod_acordo)

def aplicar_criterios_desempate(registros, regra_aplicada, tem_cod_acordo):
    """
    Aplica critérios de desempate quando múltiplos registros atendem à mesma regra
    Critérios em ordem:
    1. Menor cod_prestacao
    2. Se empate: menor cod_acordo (se existir)
    3. Se empate: primeiro registro (index menor)
    """
    if len(registros) == 1:
        return registros.iloc[[0]]
    
    print(f"         🎲 Desempate necessário para {regra_aplicada}: {len(registros)} registros")
    
    # Critério 1: Menor cod_prestacao
    menor_cod_prestacao = registros['cod_prestacao'].min()
    candidatos = registros[registros['cod_prestacao'] == menor_cod_prestacao]
    
    if len(candidatos) == 1:
        print(f"         ✅ Desempate por menor cod_prestacao: {menor_cod_prestacao}")
        return candidatos.iloc[[0]]
    
    print(f"         🎲 Empate em cod_prestacao ({menor_cod_prestacao}), aplicando critério 2")
    
    # Critério 2: Menor cod_acordo (se existir)
    if tem_cod_acordo and 'cod_acordo' in candidatos.columns:
        if 'cod_acordo_norm' not in candidatos.columns:
            candidatos = candidatos.copy()
            candidatos['cod_acordo_norm'] = pd.to_numeric(candidatos['cod_acordo'], errors='coerce').fillna(999999)
        
        menor_cod_acordo = candidatos['cod_acordo_norm'].min()
        candidatos_final = candidatos[candidatos['cod_acordo_norm'] == menor_cod_acordo]
        
        if len(candidatos_final) == 1:
            print(f"         ✅ Desempate por menor cod_acordo: {menor_cod_acordo}")
            return candidatos_final.iloc[[0]]
        
        candidatos = candidatos_final
        print(f"         🎲 Empate em cod_acordo ({menor_cod_acordo}), aplicando critério 3")
    
    # Critério 3: Primeiro registro (menor index)
    print(f"         ✅ Desempate por primeiro registro (index menor)")
    registro_escolhido = candidatos.iloc[[0]]
    
    return registro_escolhido

def salvar_arquivo_com_formatacao(df, caminho_arquivo):
    """
    Salva os registros CORRETOS escolhidos pelas regras de duplicatas
    """
    # Com a nova lógica, todos os registros do df são considerados corretos
    df_registros_corretos = df.copy()
    
    print(f"   📝 Salvando registros corretos escolhidos: {len(df_registros_corretos)} registros")
    
    # Verificar se há registros para salvar
    if len(df_registros_corretos) == 0:
        print(f"   ⚠️ Nenhum registro para salvar")
        # Criar arquivo vazio com cabeçalhos apenas
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros_Corretos"
        
        # Adicionar apenas cabeçalhos
        headers = ['Nenhum registro encontrado']
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
            ws.cell(row=1, column=col_num).font = Font(bold=True)
        
        wb.save(caminho_arquivo)
        return
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros_Corretos"
    
    # Adicionar cabeçalhos (remover coluna auxiliar se existir)
    headers = list(df_registros_corretos.columns)
    if 'eh_menor_cod' in headers:
        headers.remove('eh_menor_cod')  # Remover coluna auxiliar do arquivo final
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        # Formatar cabeçalho
        ws.cell(row=1, column=col_num).font = Font(bold=True)
    
    # Definir estilo verde para destacar registros corretos escolhidos
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    green_font = Font(color="006400", bold=True)
    
    # Adicionar os dados dos registros corretos
    for row_num, (index, row) in enumerate(df_registros_corretos.iterrows(), 2):
        for col_num, header in enumerate(headers, 1):
            cell_value = row[header]
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Aplicar formatação verde em todos os registros corretos
            cell.fill = green_fill
            cell.font = green_font
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar arquivo
    wb.save(caminho_arquivo)

def processar_filtro_duplicatas(caminho_arquivo, caminho_salvar, progresso_var, progresso_label, status_label, botao_iniciar, botao_parar, botao_cancelar, botao_arquivo):
    """Processa o arquivo removendo duplicatas"""
    global linhas_processadas, log_duplicatas, total_duplicatas
    
    try:
        # Resetar contadores
        linhas_processadas = 0
        total_duplicatas = 0
        log_duplicatas = []
        
        # Atualizar interface
        status_label.config(text="📂 Carregando arquivo...")
        progresso_var.set(10)
        
        # Carregar arquivo
        print(f"📂 Carregando arquivo: {caminho_arquivo}")
        df = pd.read_excel(caminho_arquivo)
        
        total_inicial = len(df)
        print(f"📊 Total de registros carregados: {total_inicial}")
        
        # Atualizar interface
        status_label.config(text="🔍 Analisando duplicatas...")
        progresso_var.set(30)
        
        # Verificar se arquivo tem dados
        if total_inicial == 0:
            raise ValueError("❌ Arquivo está vazio")
        
        # Filtrar duplicatas
        start_time = time.time()
        df_duplicados = filtrar_duplicatas_cpf_data(df)
        
        linhas_processadas = len(df_duplicados)
        
        # Verificar se há duplicados para processar
        if linhas_processadas == 0:
            raise ValueError("❌ Nenhum duplicado encontrado no arquivo")
        
        # Atualizar progresso
        progresso_var.set(70)
        status_label.config(text="💾 Salvando registros corretos...")
        
        # Salvar arquivo com formatação (registros corretos escolhidos)
        salvar_arquivo_com_formatacao(df_duplicados, caminho_salvar)
        print(f"💾 Arquivo salvo: {caminho_salvar}")
        
        # Contar registros salvos (todos são corretos com a nova lógica)
        registros_corretos_salvos = linhas_processadas
        duplicatas_resolvidas = len(df_duplicados.groupby(['cpf', 'data_vencimento', 'numero_prestacao']))
        
        # Atualizar progresso
        progresso_var.set(90)
        
        # Salvar relatório de duplicatas
        if log_duplicatas:
            relatorio_file = caminho_salvar.replace('.xlsx', '_relatorio_duplicatas.txt')
            with open(relatorio_file, "w", encoding="utf-8") as f:
                f.write(f"=== RELATÓRIO DE RESOLUÇÃO DE DUPLICATAS ===\n")
                f.write(f"Data/Hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Arquivo original: {os.path.basename(caminho_arquivo)}\n")
                f.write(f"Arquivo de resultado: {os.path.basename(caminho_salvar)}\n")
                f.write(f"Total de registros originais: {total_inicial}\n")
                f.write(f"Grupos de duplicatas encontrados: {duplicatas_resolvidas}\n")
                f.write(f"Registros corretos salvos: {registros_corretos_salvos}\n")
                f.write(f"Taxa de duplicação: {(duplicatas_resolvidas/total_inicial*100) if total_inicial > 0 else 0:.1f}%\n")
                f.write("=" * 60 + "\n\n")
                f.write("REGRAS APLICADAS:\n")
                f.write("1. Se tiver data_pagamento: salva linha sem data_pagamento\n")
                f.write("2. Se ambos têm data_pagamento: prioriza cod_acordo=0\n")
                f.write("3. Se ambos têm data_pagamento e cod_acordo: menor cod_prestacao\n")
                f.write("4. Se data_pagamento null: prioriza cod_acordo=0\n")
                f.write("5. Se data_pagamento null e ambos têm cod_acordo: menor cod_prestacao\n")
                f.write("=" * 60 + "\n\n")
                f.write("GRUPOS PROCESSADOS:\n")
                f.write("-" * 30 + "\n")
                for linha in log_duplicatas:
                    f.write(linha + "\n")
            print(f"📝 Relatório de duplicatas salvo: {relatorio_file}")
        
        # Finalizar
        elapsed_time = time.time() - start_time
        progresso_var.set(100)
        
        # Mensagem de sucesso
        final_msg = f"✅ Análise concluída! {registros_corretos_salvos} registros corretos de {duplicatas_resolvidas} grupos em {elapsed_time:.1f}s"
        status_label.config(text=final_msg)
        print(final_msg)
        
        # Mostrar resultado em popup
        messagebox.showinfo("Resolução de Duplicatas Concluída", 
                           f"✅ Processo concluído com sucesso!\n\n"
                           f"📊 Registros originais: {total_inicial}\n"
                           f"🔍 Grupos de duplicatas: {duplicatas_resolvidas}\n"
                           f"✅ Registros corretos salvos: {registros_corretos_salvos}\n"
                           f"⏱️ Tempo: {elapsed_time:.1f}s\n\n"
                           f"📁 Arquivo salvo: {os.path.basename(caminho_salvar)}")
        
    except Exception as e:
        error_msg = f"❌ Erro no processamento: {str(e)}"
        status_label.config(text=error_msg)
        print(error_msg)
        messagebox.showerror("Erro", error_msg)
    
    finally:
        # Reativar botões
        botao_iniciar.config(state=tk.NORMAL)
        botao_arquivo.config(state=tk.NORMAL)
        botao_parar.config(state=tk.DISABLED)
        botao_cancelar.config(state=tk.DISABLED)
        progresso_label.config(text="")

def executar_filtro_duplicatas(caminho_arquivo, caminho_salvar, progresso_var, progresso_label, status_label, botao_iniciar, botao_parar, botao_cancelar, botao_arquivo):
    """Executa o processamento em thread separada"""
    try:
        # Resetar flag de parada
        parar_flag.clear()
        
        # Desativar botões durante processamento
        botao_iniciar.config(state=tk.DISABLED)
        botao_arquivo.config(state=tk.DISABLED)
        botao_parar.config(state=tk.NORMAL)
        botao_cancelar.config(state=tk.NORMAL)
        
        # Executar processamento em thread separada
        thread = threading.Thread(
            target=processar_filtro_duplicatas,
            args=(caminho_arquivo, caminho_salvar, progresso_var, progresso_label, status_label, botao_iniciar, botao_parar, botao_cancelar, botao_arquivo)
        )
        thread.daemon = True
        thread.start()
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao iniciar processamento: {str(e)}")
        # Reativar botões em caso de erro
        botao_iniciar.config(state=tk.NORMAL)
        botao_arquivo.config(state=tk.NORMAL)
        botao_parar.config(state=tk.DISABLED)
        botao_cancelar.config(state=tk.DISABLED)
